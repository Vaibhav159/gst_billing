from datetime import datetime, timedelta

from django.urls import reverse
from rest_framework import status

from billing.constants import INVOICE_TYPE_INWARD, INVOICE_TYPE_OUTWARD
from billing.models import Business, Customer, Invoice, LineItem
from billing.tests.test_base import BaseAPITestCase


class ReportViewTestCase(BaseAPITestCase):
    def setUp(self):
        # Call parent setUp to set up authentication
        super().setUp()

        # Create test data
        self.business = Business.objects.create(
            name="Test Business Report",
            address="Test Address",
            gst_number="27AADCB2230M1Z3",
            mobile_number="1234567890",
            state_name="MAHARASHTRA",
        )

        self.customer = Customer.objects.create(
            name="Test Customer Report",
            address="Test Customer Address",
            gst_number="27AADCB2230M1Z3",
            mobile_number="9876543210",
            state_name="MAHARASHTRA",
        )
        self.customer.businesses.add(self.business)

        # Create an outward invoice
        self.outward_invoice = Invoice.objects.create(
            customer=self.customer,
            business=self.business,
            invoice_number="OUT001",
            invoice_date=datetime.now().date(),
            type_of_invoice=INVOICE_TYPE_OUTWARD,
        )

        # Create an inward invoice
        self.inward_invoice = Invoice.objects.create(
            customer=self.customer,
            business=self.business,
            invoice_number="IN001",
            invoice_date=datetime.now().date(),
            type_of_invoice=INVOICE_TYPE_INWARD,
        )

        # Create line items for both invoices
        LineItem.create_line_item_for_invoice(
            product_name="Test Product",
            quantity=10,
            rate=100,
            invoice_id=self.outward_invoice.id,
        )

        LineItem.create_line_item_for_invoice(
            product_name="Test Product",
            quantity=5,
            rate=100,
            invoice_id=self.inward_invoice.id,
        )

    def test_report_generation(self):
        """Test that reports are generated correctly with proper invoice type filtering"""
        url = reverse("generate-report")

        # Set date range to include our test invoices
        today = datetime.now().date()
        start_date = (today - timedelta(days=7)).strftime("%Y-%m-%d")
        end_date = (today + timedelta(days=1)).strftime("%Y-%m-%d")

        # Test outward report
        data = {
            "start_date": start_date,
            "end_date": end_date,
            "invoice_type": INVOICE_TYPE_OUTWARD,
        }

        response = self.client.post(url, data, format="json")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        # Test inward report
        data["invoice_type"] = INVOICE_TYPE_INWARD
        response = self.client.post(url, data, format="json")
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        # Test both types
        data["invoice_type"] = "both"
        response = self.client.post(url, data, format="json")
        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_line_item_data_for_download(self):
        """Test that line item data includes invoice type and ID"""
        today = datetime.now().date()
        start_date = (today - timedelta(days=7)).strftime("%Y-%m-%d")
        end_date = (today + timedelta(days=1)).strftime("%Y-%m-%d")

        # Get line item data
        line_items = LineItem.get_line_item_data_for_download(
            start_date=start_date,
            end_date=end_date,
            business=self.business,
        )

        # Check that we have data for both invoices
        self.assertEqual(len(line_items), 2)

        # Check that each line item has the correct invoice type
        outward_items = [
            item for item in line_items if item[14] == INVOICE_TYPE_OUTWARD
        ]
        inward_items = [item for item in line_items if item[14] == INVOICE_TYPE_INWARD]

        self.assertEqual(len(outward_items), 1)
        self.assertEqual(len(inward_items), 1)

        # Verify the invoice IDs are included
        self.assertTrue(any(item[15] == self.outward_invoice.id for item in line_items))
        self.assertTrue(any(item[15] == self.inward_invoice.id for item in line_items))


    def test_both_report_net_row_subtracts_inward_from_outward(self):
        """The combined row is NET (outward − inward) — it used to ADD sales
        to purchases, which the CA review flagged as meaningless (D7)."""
        import io

        from openpyxl import load_workbook

        today = datetime.now().date()
        data = {
            "start_date": (today - timedelta(days=7)).strftime("%Y-%m-%d"),
            "end_date": (today + timedelta(days=1)).strftime("%Y-%m-%d"),
            "invoice_type": "both",
        }
        response = self.client.post(reverse("generate-report"), data, format="json")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        wb = load_workbook(io.BytesIO(response.content), data_only=True)
        sheet = wb["Test Business Report"]

        net_rows = [
            row for row in sheet.iter_rows(values_only=True)
            if any(isinstance(c, str) and c.startswith("NET (OUTWARD") for c in row)
        ]
        self.assertEqual(len(net_rows), 1, "exactly one NET row expected")
        row = net_rows[0]
        # Fixture: outward 10×100, inward 5×100 → net taxable 500.
        self.assertEqual(float(row[10]), 500.0)
        # And no additive GRAND TOTAL row survives anywhere.
        for r in sheet.iter_rows(values_only=True):
            for c in r:
                if isinstance(c, str):
                    self.assertFalse(c.startswith("GRAND TOTAL"), f"additive row still present: {c}")
